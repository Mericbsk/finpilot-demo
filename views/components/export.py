# -*- coding: utf-8 -*-
"""
FinPilot Export Component
=========================

Scan sonuçlarını CSV, Excel ve PDF formatlarında dışa aktarma.

Usage:
    from views.components.export import render_export_panel, export_to_excel

    render_export_panel(df)
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter


def _sanitize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sanitize DataFrame column names for Excel compatibility.
    Removes/replaces special characters that Excel doesn't support.
    """

    def clean_name(name: str) -> str:
        if not isinstance(name, str):
            name = str(name)
        # Replace brackets and special chars with underscores
        name = re.sub(r"[\[\]\(\)\{\}]", "", name)
        # Replace other problematic characters
        name = re.sub(r"[^\w\s\-_]", "_", name)
        # Remove leading/trailing whitespace
        name = name.strip()
        # Ensure it's not empty
        return name if name else "Column"

    df = df.copy()
    df.columns = [clean_name(col) for col in df.columns]
    return df


# ============================================
# 📊 Excel Export
# ============================================


def export_to_excel(
    df: pd.DataFrame, filename: Optional[str] = None, sheet_name: str = "Scan Results"
) -> bytes:
    """
    Export DataFrame to Excel format with styling.

    Args:
        df: DataFrame to export
        filename: Optional filename (not used, kept for API compatibility)
        sheet_name: Excel sheet name

    Returns:
        Excel file as bytes
    """
    output = io.BytesIO()

    # Sanitize column names for Excel compatibility
    df = _sanitize_column_names(df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Get workbook and worksheet for styling
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Auto-adjust column widths (use get_column_letter for >26 columns)
        for idx, col in enumerate(df.columns, start=1):
            max_length = (
                max(df[col].astype(str).str.len().max() if not df.empty else 0, len(str(col))) + 2
            )
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(max_length, 40)

        # Style header row
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Highlight buyable rows
        buy_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        if "entry_ok" in df.columns:
            for row_idx, entry_ok in enumerate(df["entry_ok"], start=2):
                if entry_ok:
                    for cell in worksheet[row_idx]:
                        cell.fill = buy_fill

    output.seek(0)
    return output.getvalue()


# ============================================
# 📄 PDF Export
# ============================================


def export_to_pdf(
    df: pd.DataFrame, title: str = "FinPilot Scan Report", include_summary: bool = True
) -> bytes:
    """
    Export DataFrame to PDF format.

    Args:
        df: DataFrame to export
        title: Report title
        include_summary: Whether to include summary statistics

    Returns:
        PDF file as bytes
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        # Fallback: generate simple HTML-based PDF
        return _export_to_html_pdf(df, title)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=12,
        textColor=colors.HexColor("#1E3A5F"),
    )
    elements.append(Paragraph(title, title_style))
    elements.append(
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
    )
    elements.append(Spacer(1, 12))

    # Summary
    if include_summary and not df.empty:
        summary_text = f"""
        <b>Summary:</b><br/>
        Total Symbols: {len(df)}<br/>
        {'Buyable: ' + str(len(df[df['entry_ok']])) if 'entry_ok' in df.columns else ''}<br/>
        {'Avg Score: ' + f"{df['recommendation_score'].mean():.1f}" if 'recommendation_score' in df.columns else ''}
        """
        elements.append(Paragraph(summary_text, styles["Normal"]))
        elements.append(Spacer(1, 12))

    # Select key columns for PDF (limit to fit page)
    key_columns = [
        "symbol",
        "price",
        "entry_ok",
        "recommendation_score",
        "take_profit",
        "stop_loss",
        "regime",
        "timestamp",
    ]
    available_cols = [c for c in key_columns if c in df.columns]

    if not available_cols:
        available_cols = df.columns[:8].tolist()  # Fallback to first 8 columns

    # Prepare table data
    table_data = [available_cols]  # Header
    for _, row in df[available_cols].head(50).iterrows():  # Limit to 50 rows
        row_data = []
        for col in available_cols:
            val = row[col]
            if isinstance(val, float):
                val = f"{val:.2f}"
            elif isinstance(val, bool):
                val = "✓" if val else ""
            else:
                val = str(val)[:20]  # Truncate long strings
            row_data.append(val)
        table_data.append(row_data)

    # Create table
    col_widths = [1.2 * inch if c == "symbol" else 0.9 * inch for c in available_cols]
    table = Table(table_data, colWidths=col_widths)

    # Table styling
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#1E293B")),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ]
    )

    # Highlight buyable rows
    if "entry_ok" in available_cols:
        entry_ok_idx = available_cols.index("entry_ok")
        for row_idx, row in enumerate(table_data[1:], start=1):
            if row[entry_ok_idx] == "✓":
                table_style.add(
                    "BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#D1FAE5")
                )

    table.setStyle(table_style)
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 24))
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"], fontSize=8, textColor=colors.gray
    )
    elements.append(
        Paragraph(
            "FinPilot Trading System - This report is for informational purposes only.",
            footer_style,
        )
    )

    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def _export_to_html_pdf(df: pd.DataFrame, title: str) -> bytes:
    """Fallback HTML-based PDF generation."""
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>{title}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1 {{ color: #1E3A5F; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th {{ background: #1E3A5F; color: white; padding: 10px; }}
            td {{ border: 1px solid #CBD5E1; padding: 8px; }}
            tr:nth-child(even) {{ background: #F1F5F9; }}
            .buyable {{ background: #D1FAE5 !important; }}
            .footer {{ margin-top: 20px; font-size: 12px; color: #666; }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
        <p><strong>Total Symbols:</strong> {len(df)}</p>
        {df.head(50).to_html(index=False, classes='data-table')}
        <p class="footer">FinPilot Trading System - This report is for informational purposes only.</p>
    </body>
    </html>
    """
    return html.encode("utf-8")


# ============================================
# 📊 CSV Export (Enhanced)
# ============================================


def export_to_csv(df: pd.DataFrame, include_metadata: bool = True) -> str:
    """
    Export DataFrame to CSV format.

    Args:
        df: DataFrame to export
        include_metadata: Whether to include metadata as comments

    Returns:
        CSV content as string
    """
    output = io.StringIO()

    if include_metadata:
        output.write("# FinPilot Scan Report\n")
        output.write(f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
        output.write(f"# Total Symbols: {len(df)}\n")
        if "entry_ok" in df.columns:
            output.write(f"# Buyable Symbols: {len(df[df['entry_ok']])}\n")
        output.write("#\n")

    df.to_csv(output, index=False)
    return output.getvalue()


# ============================================
# 🎨 Export UI Component
# ============================================


def render_export_panel(df: pd.DataFrame) -> None:
    """
    Render the export panel with format selection and download buttons.

    Args:
        df: DataFrame to export
    """
    if df is None or df.empty:
        st.info("📭 Dışa aktarılacak veri yok. Önce bir tarama yapın.")
        return

    st.markdown("### 📥 Sonuçları Dışa Aktar")

    # Format selection
    col1, col2, col3 = st.columns(3)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    # CSV Export
    with col1:
        csv_data = export_to_csv(df)
        st.download_button(
            label="📄 CSV İndir",
            data=csv_data,
            file_name=f"finpilot_scan_{timestamp}.csv",
            mime="text/csv",
            use_container_width=True,
            key="export_csv",
        )

    # Excel Export
    with col2:
        try:
            excel_data = export_to_excel(df)
            st.download_button(
                label="📊 Excel İndir",
                data=excel_data,
                file_name=f"finpilot_scan_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="export_excel",
            )
        except ImportError:
            st.button(
                "📊 Excel (openpyxl gerekli)",
                disabled=True,
                use_container_width=True,
                key="export_excel_disabled",
            )

    # PDF Export
    with col3:
        try:
            pdf_data = export_to_pdf(df)
            if pdf_data.startswith(b"<!DOCTYPE"):
                # HTML fallback
                st.download_button(
                    label="📄 HTML Rapor",
                    data=pdf_data,
                    file_name=f"finpilot_scan_{timestamp}.html",
                    mime="text/html",
                    use_container_width=True,
                    key="export_html",
                )
            else:
                st.download_button(
                    label="📑 PDF İndir",
                    data=pdf_data,
                    file_name=f"finpilot_scan_{timestamp}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="export_pdf",
                )
        except Exception:
            st.button(
                "📑 PDF (reportlab gerekli)",
                disabled=True,
                use_container_width=True,
                key="export_pdf_disabled",
            )

    # Export options
    with st.expander("⚙️ Dışa Aktarma Seçenekleri", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            include_all = st.checkbox("Tüm sütunları dahil et", value=True, key="export_all_cols")
            buyable_only = st.checkbox(
                "Sadece alınabilir sembolleri dahil et", value=False, key="export_buyable_only"
            )
        with col2:
            include_metadata = st.checkbox("Meta veri ekle", value=True, key="export_metadata")
            top_n = st.number_input(
                "Maksimum satır", min_value=10, max_value=500, value=100, key="export_max_rows"
            )

        # Apply filters for custom export
        export_df = df.copy()
        if buyable_only and "entry_ok" in export_df.columns:
            export_df = export_df[export_df["entry_ok"]]
        export_df = export_df.head(int(top_n))

        if not include_all:
            key_cols = [
                "symbol",
                "price",
                "entry_ok",
                "recommendation_score",
                "take_profit",
                "stop_loss",
                "regime",
            ]
            export_df = export_df[[c for c in key_cols if c in export_df.columns]]

        st.caption(f"📊 Dışa aktarılacak: {len(export_df)} satır, {len(export_df.columns)} sütun")

        # Custom export button
        custom_csv = export_to_csv(export_df, include_metadata=include_metadata)
        st.download_button(
            label="📥 Özel CSV İndir",
            data=custom_csv,
            file_name=f"finpilot_custom_{timestamp}.csv",
            mime="text/csv",
            key="export_custom_csv",
        )


def render_export_button_row(df: pd.DataFrame) -> None:
    """
    Render a compact row of export buttons for inline use.

    Args:
        df: DataFrame to export
    """
    if df is None or df.empty:
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    col1, col2, col3 = st.columns(3)

    with col1:
        csv_data = export_to_csv(df, include_metadata=False)
        st.download_button(
            "📄 CSV",
            data=csv_data,
            file_name=f"scan_{timestamp}.csv",
            mime="text/csv",
            key=f"quick_csv_{timestamp}",
        )

    with col2:
        try:
            excel_data = export_to_excel(df)
            st.download_button(
                "📊 Excel",
                data=excel_data,
                file_name=f"scan_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"quick_excel_{timestamp}",
            )
        except ImportError:
            st.caption("Excel N/A")

    with col3:
        try:
            pdf_data = export_to_pdf(df)
            mime_type = "text/html" if pdf_data.startswith(b"<!DOCTYPE") else "application/pdf"
            ext = "html" if pdf_data.startswith(b"<!DOCTYPE") else "pdf"
            st.download_button(
                "📑 PDF",
                data=pdf_data,
                file_name=f"scan_{timestamp}.{ext}",
                mime=mime_type,
                key=f"quick_pdf_{timestamp}",
            )
        except Exception:
            st.caption("PDF N/A")


__all__ = [
    "export_to_csv",
    "export_to_excel",
    "export_to_pdf",
    "render_export_panel",
    "render_export_button_row",
]
